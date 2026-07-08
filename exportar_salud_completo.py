
import argparse
import csv
import json
import re
import sqlite3
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook


EXCEL_MAX_ROWS = 1_048_576
EXCEL_DATA_ROWS = EXCEL_MAX_ROWS - 1
VERSION = "2026-07-08_formato_excel_v3_ruta_carpeta"
INSERT_CHUNK = 10_000


TIPOS_ES = {
    "HKQuantityTypeIdentifierStepCount": "Pasos",
    "HKQuantityTypeIdentifierHeartRate": "Frecuencia cardiaca",
    "HKQuantityTypeIdentifierRestingHeartRate": "Frecuencia cardiaca en reposo",
    "HKQuantityTypeIdentifierWalkingHeartRateAverage": "Frecuencia cardiaca caminando",
    "HKQuantityTypeIdentifierHeartRateVariabilitySDNN": "Variabilidad cardiaca HRV",
    "HKQuantityTypeIdentifierHeartRateRecoveryOneMinute": "Recuperación cardiaca 1 min",
    "HKQuantityTypeIdentifierBodyMass": "Peso",
    "HKQuantityTypeIdentifierHeight": "Altura",
    "HKQuantityTypeIdentifierBodyMassIndex": "IMC",
    "HKQuantityTypeIdentifierActiveEnergyBurned": "Energía activa",
    "HKQuantityTypeIdentifierBasalEnergyBurned": "Energía basal",
    "HKQuantityTypeIdentifierDistanceWalkingRunning": "Distancia andando/corriendo",
    "HKQuantityTypeIdentifierDistanceCycling": "Distancia ciclismo",
    "HKQuantityTypeIdentifierFlightsClimbed": "Pisos subidos",
    "HKQuantityTypeIdentifierAppleExerciseTime": "Tiempo de ejercicio",
    "HKQuantityTypeIdentifierAppleStandTime": "Tiempo de pie",
    "HKQuantityTypeIdentifierOxygenSaturation": "Saturación de oxígeno",
    "HKQuantityTypeIdentifierRespiratoryRate": "Frecuencia respiratoria",
    "HKQuantityTypeIdentifierBodyTemperature": "Temperatura corporal",
    "HKQuantityTypeIdentifierVO2Max": "VO2 máx.",
    "HKQuantityTypeIdentifierEnvironmentalAudioExposure": "Ruido ambiental",
    "HKQuantityTypeIdentifierHeadphoneAudioExposure": "Ruido auriculares",
    "HKCategoryTypeIdentifierSleepAnalysis": "Sueño",
    "HKCategoryTypeIdentifierAppleStandHour": "Hora de pie",
    "HKCategoryTypeIdentifierMindfulSession": "Mindfulness",
    "HKActivitySummary": "Resumen actividad diaria",
}

SLEEP_VALUES_ES = {
    "HKCategoryValueSleepAnalysisInBed": "En la cama",
    "HKCategoryValueSleepAnalysisAsleep": "Dormido",
    "HKCategoryValueSleepAnalysisAsleepUnspecified": "Dormido sin especificar",
    "HKCategoryValueSleepAnalysisAsleepCore": "Sueño ligero/Core",
    "HKCategoryValueSleepAnalysisAsleepDeep": "Sueño profundo",
    "HKCategoryValueSleepAnalysisAsleepREM": "Sueño REM",
    "HKCategoryValueSleepAnalysisAwake": "Despierto",
}

CATEGORY_VALUES_ES = {
    **SLEEP_VALUES_ES,
    "HKCategoryValueAppleStandHourStood": "De pie",
    "HKCategoryValueAppleStandHourIdle": "Sin ponerse de pie",
}

WORKOUTS_ES = {
    "HKWorkoutActivityTypeWalking": "Entreno caminar",
    "HKWorkoutActivityTypeRunning": "Entreno correr",
    "HKWorkoutActivityTypeCycling": "Entreno ciclismo",
    "HKWorkoutActivityTypeTraditionalStrengthTraining": "Entreno fuerza tradicional",
    "HKWorkoutActivityTypeFunctionalStrengthTraining": "Entreno fuerza funcional",
    "HKWorkoutActivityTypeSwimming": "Entreno natación",
    "HKWorkoutActivityTypeYoga": "Entreno yoga",
    "HKWorkoutActivityTypeOther": "Entreno otro",
    "HKWorkoutActivityTypeMixedCardio": "Entreno cardio mixto",
    "HKWorkoutActivityTypeHighIntensityIntervalTraining": "Entreno HIIT",
}


def limpiar_tipo_tecnico(tipo):
    if not tipo:
        return "Sin tipo"
    for prefix in [
        "HKQuantityTypeIdentifier",
        "HKCategoryTypeIdentifier",
        "HKWorkoutActivityType",
        "HKDataType",
        "HK",
    ]:
        if tipo.startswith(prefix):
            tipo = tipo[len(prefix):]
            break
    # CamelCase -> palabras
    tipo = re.sub(r"(?<!^)(?=[A-Z])", " ", tipo).strip()
    return tipo or "Sin tipo"


def nombre_tipo(tipo_original):
    if tipo_original in TIPOS_ES:
        return TIPOS_ES[tipo_original]
    if tipo_original in WORKOUTS_ES:
        return WORKOUTS_ES[tipo_original]
    return limpiar_tipo_tecnico(tipo_original)


def parse_fecha_apple(fecha):
    if not fecha:
        return None
    formatos = [
        "%Y-%m-%d %H:%M:%S %z",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]
    for fmt in formatos:
        try:
            dt = datetime.strptime(fecha, fmt)
            # Conservamos la hora local que aparece en el export.
            return dt.replace(tzinfo=None)
        except ValueError:
            pass
    return None


def fecha_a_texto(dt):
    if not dt:
        return ""
    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def parse_fecha_arg(valor, nombre):
    if not valor:
        return None
    try:
        return datetime.strptime(valor, "%Y-%m-%d")
    except ValueError:
        raise SystemExit(f"Fecha inválida en {nombre}: {valor}. Usa formato YYYY-MM-DD, ejemplo 2024-01-31.")


def normalizar_detalle(detalle):
    d = (detalle or "hora").strip().lower()
    aliases = {
        "dia": "dia",
        "día": "dia",
        "day": "dia",
        "hora": "hora",
        "hour": "hora",
        "min": "min",
        "minuto": "min",
        "minute": "min",
        "todos": "todos",
        "todo": "todos",
        "all": "todos",
        "raw": "todos",
    }
    if d not in aliases:
        raise SystemExit("Detalle inválido. Usa uno de estos: dia, hora, min, todos.")
    return aliases[d]


def buscar_export_xml(xml_arg):
    """
    Devuelve la ruta real a export.xml.

    Acepta dos casos:
    - --xml apuntando directamente a ...\\export.xml
    - --xml apuntando a una carpeta; en ese caso busca export.xml dentro
      de esa carpeta y sus subcarpetas.
    """
    if xml_arg:
        p = Path(xml_arg)

        if not p.exists():
            raise SystemExit(f"No existe la ruta indicada: {p}")

        if p.is_file():
            if p.name.lower() != "export.xml":
                raise SystemExit(
                    f"La ruta indicada es un archivo, pero no se llama export.xml: {p}\n"
                    "Indica el archivo export.xml de Apple Salud o una carpeta que lo contenga."
                )
            return p

        if p.is_dir():
            encontrados = sorted(p.rglob("export.xml"), key=lambda x: (len(x.parts), str(x).lower()))
            if encontrados:
                return encontrados[0]

            otros_xml = sorted(p.rglob("*.xml"), key=lambda x: str(x).lower())[:20]
            msg = [
                f"La ruta indicada es una carpeta, pero no he encontrado export.xml dentro: {p}",
            ]
            if otros_xml:
                msg.append("XML encontrados en esa carpeta:")
                msg.extend(f"- {x}" for x in otros_xml)
            msg.append("Indica la carpeta correcta o la ruta completa al archivo export.xml.")
            raise SystemExit("\n".join(msg))

        raise SystemExit(f"La ruta indicada no es un archivo ni una carpeta válida: {p}")

    encontrados = sorted(Path(".").rglob("export.xml"), key=lambda x: (len(x.parts), str(x).lower()))
    if not encontrados:
        raise SystemExit("No he encontrado export.xml en esta carpeta ni en subcarpetas.")
    return encontrados[0]

def dentro_de_fechas(inicio_dt, fecha_inicio, fecha_fin_exclusiva):
    if inicio_dt is None:
        return not fecha_inicio and not fecha_fin_exclusiva
    if fecha_inicio and inicio_dt < fecha_inicio:
        return False
    if fecha_fin_exclusiva and inicio_dt >= fecha_fin_exclusiva:
        return False
    return True


def a_float(valor):
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except Exception:
        return None


def duracion_minutos(inicio_dt, fin_dt):
    if not inicio_dt or not fin_dt:
        return None
    return round((fin_dt - inicio_dt).total_seconds() / 60, 6)


def normalizar_unidad_y_valor(tipo_original, valor_original, unidad_original):
    valor = a_float(valor_original)
    unidad = unidad_original or ""

    if valor is None:
        return None, valor_original or "", unidad

    # Frecuencia cardiaca: Apple suele exportarla como count/min.
    if tipo_original == "HKQuantityTypeIdentifierHeartRate":
        return round(valor, 3), str(round(valor, 3)), "lpm"

    if tipo_original in {
        "HKQuantityTypeIdentifierRestingHeartRate",
        "HKQuantityTypeIdentifierWalkingHeartRateAverage",
        "HKQuantityTypeIdentifierHeartRateRecoveryOneMinute",
    }:
        return round(valor, 3), str(round(valor, 3)), "lpm"

    if tipo_original == "HKQuantityTypeIdentifierHeartRateVariabilitySDNN":
        return round(valor, 3), str(round(valor, 3)), "ms"

    if tipo_original == "HKQuantityTypeIdentifierStepCount":
        return round(valor, 3), str(round(valor, 3)), "pasos"

    if tipo_original == "HKQuantityTypeIdentifierBodyMass":
        return round(valor, 3), str(round(valor, 3)), "kg"

    if tipo_original == "HKQuantityTypeIdentifierHeight":
        if unidad_original == "m":
            valor *= 100
        return round(valor, 3), str(round(valor, 3)), "cm"

    if tipo_original in {
        "HKQuantityTypeIdentifierActiveEnergyBurned",
        "HKQuantityTypeIdentifierBasalEnergyBurned",
    }:
        return round(valor, 3), str(round(valor, 3)), "kcal"

    if tipo_original in {
        "HKQuantityTypeIdentifierDistanceWalkingRunning",
        "HKQuantityTypeIdentifierDistanceCycling",
    }:
        if unidad_original in {"m", "meter", "meters"}:
            valor /= 1000
        elif unidad_original in {"mi", "mile", "miles"}:
            valor *= 1.609344
        return round(valor, 6), str(round(valor, 6)), "km"

    if tipo_original in {
        "HKQuantityTypeIdentifierAppleExerciseTime",
        "HKQuantityTypeIdentifierAppleStandTime",
    }:
        return round(valor, 3), str(round(valor, 3)), "min"

    if tipo_original == "HKQuantityTypeIdentifierOxygenSaturation":
        # Suposición explícita: Apple puede exportar 0.97 para 97%.
        if valor <= 1:
            valor *= 100
        return round(valor, 3), str(round(valor, 3)), "%"

    if tipo_original == "HKQuantityTypeIdentifierRespiratoryRate":
        return round(valor, 3), str(round(valor, 3)), "resp/min"

    if tipo_original == "HKQuantityTypeIdentifierBodyTemperature":
        return round(valor, 3), str(round(valor, 3)), "°C"

    if tipo_original == "HKQuantityTypeIdentifierVO2Max":
        return round(valor, 3), str(round(valor, 3)), "ml/kg/min"

    return round(valor, 6), str(round(valor, 6)), unidad


def convertir_distancia_a_km(valor, unidad):
    v = a_float(valor)
    if v is None:
        return None
    if unidad in {"km", "kilometer", "kilometers"}:
        return v
    if unidad in {"m", "meter", "meters"}:
        return v / 1000
    if unidad in {"mi", "mile", "miles"}:
        return v * 1.609344
    return None


def convertir_energia_a_kcal(valor, unidad):
    v = a_float(valor)
    if v is None:
        return None
    if unidad in {"kcal", "Cal", "calorie", "calories", "Calorie", "Calories"}:
        return v
    if unidad in {"kJ", "kilojoule", "kilojoules"}:
        return v / 4.184
    return None


def convertir_duracion_a_min(valor, unidad):
    v = a_float(valor)
    if v is None:
        return None
    if unidad in {"min", "minute", "minutes"}:
        return v
    if unidad in {"s", "sec", "second", "seconds"}:
        return v / 60
    if unidad in {"h", "hr", "hour", "hours"}:
        return v * 60
    return v


def crear_db(db_path):
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE salud (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clase TEXT,
            tipo_original TEXT,
            tipo TEXT,
            fuente TEXT,
            unidad_original TEXT,
            unidad TEXT,
            valor_original TEXT,
            valor_num REAL,
            valor_texto TEXT,
            inicio TEXT,
            fin TEXT,
            duracion_min REAL,
            creacion TEXT,
            distancia_km REAL,
            energia_kcal REAL,
            extra TEXT
        )
    """)
    cur.execute("CREATE INDEX idx_tipo ON salud(tipo_original)")
    cur.execute("CREATE INDEX idx_inicio ON salud(inicio)")
    conn.commit()
    return conn


def insertar_lote(conn, lote):
    conn.executemany("""
        INSERT INTO salud (
            clase, tipo_original, tipo, fuente, unidad_original, unidad,
            valor_original, valor_num, valor_texto,
            inicio, fin, duracion_min, creacion,
            distancia_km, energia_kcal, extra
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, lote)
    conn.commit()


def procesar_xml(xml_path, conn, fecha_inicio, fecha_fin_exclusiva):
    lote = []
    contadores = defaultdict(int)

    for _, elem in ET.iterparse(xml_path, events=("end",)):
        tag = elem.tag

        if tag == "Record":
            tipo_original = elem.attrib.get("type", "")
            inicio_dt = parse_fecha_apple(elem.attrib.get("startDate", ""))
            fin_dt = parse_fecha_apple(elem.attrib.get("endDate", ""))
            creacion_dt = parse_fecha_apple(elem.attrib.get("creationDate", ""))

            if not dentro_de_fechas(inicio_dt, fecha_inicio, fecha_fin_exclusiva):
                elem.clear()
                continue

            fuente = elem.attrib.get("sourceName", "")
            valor_original = elem.attrib.get("value", "")
            unidad_original = elem.attrib.get("unit", "")

            if tipo_original.startswith("HKCategoryTypeIdentifier"):
                clase = "Categoría"
                tipo = nombre_tipo(tipo_original)
                valor_texto = CATEGORY_VALUES_ES.get(valor_original, valor_original)
                valor_num = None
                unidad = "estado"
            else:
                clase = "Registro"
                tipo = nombre_tipo(tipo_original)
                valor_num, valor_texto, unidad = normalizar_unidad_y_valor(
                    tipo_original,
                    valor_original,
                    unidad_original
                )

            lote.append((
                clase,
                tipo_original,
                tipo,
                fuente,
                unidad_original,
                unidad,
                valor_original,
                valor_num,
                valor_texto,
                fecha_a_texto(inicio_dt),
                fecha_a_texto(fin_dt),
                duracion_minutos(inicio_dt, fin_dt),
                fecha_a_texto(creacion_dt),
                None,
                None,
                ""
            ))
            contadores[clase] += 1

        elif tag == "Category":
            # Fallback por si algún export antiguo apareciera así.
            tipo_original = elem.attrib.get("type", "")
            inicio_dt = parse_fecha_apple(elem.attrib.get("startDate", ""))
            fin_dt = parse_fecha_apple(elem.attrib.get("endDate", ""))
            creacion_dt = parse_fecha_apple(elem.attrib.get("creationDate", ""))

            if not dentro_de_fechas(inicio_dt, fecha_inicio, fecha_fin_exclusiva):
                elem.clear()
                continue

            valor_original = elem.attrib.get("value", "")
            lote.append((
                "Categoría",
                tipo_original,
                nombre_tipo(tipo_original),
                elem.attrib.get("sourceName", ""),
                "",
                "estado",
                valor_original,
                None,
                CATEGORY_VALUES_ES.get(valor_original, valor_original),
                fecha_a_texto(inicio_dt),
                fecha_a_texto(fin_dt),
                duracion_minutos(inicio_dt, fin_dt),
                fecha_a_texto(creacion_dt),
                None,
                None,
                ""
            ))
            contadores["Categoría"] += 1

        elif tag == "Workout":
            tipo_original = elem.attrib.get("workoutActivityType", "")
            inicio_dt = parse_fecha_apple(elem.attrib.get("startDate", ""))
            fin_dt = parse_fecha_apple(elem.attrib.get("endDate", ""))
            creacion_dt = parse_fecha_apple(elem.attrib.get("creationDate", ""))

            if not dentro_de_fechas(inicio_dt, fecha_inicio, fecha_fin_exclusiva):
                elem.clear()
                continue

            duracion = convertir_duracion_a_min(
                elem.attrib.get("duration", ""),
                elem.attrib.get("durationUnit", "")
            )

            distancia_km = convertir_distancia_a_km(
                elem.attrib.get("totalDistance", ""),
                elem.attrib.get("totalDistanceUnit", "")
            )

            energia_kcal = convertir_energia_a_kcal(
                elem.attrib.get("totalEnergyBurned", ""),
                elem.attrib.get("totalEnergyBurnedUnit", "")
            )

            extra = {
                "duration": elem.attrib.get("duration", ""),
                "durationUnit": elem.attrib.get("durationUnit", ""),
                "totalDistance": elem.attrib.get("totalDistance", ""),
                "totalDistanceUnit": elem.attrib.get("totalDistanceUnit", ""),
                "totalEnergyBurned": elem.attrib.get("totalEnergyBurned", ""),
                "totalEnergyBurnedUnit": elem.attrib.get("totalEnergyBurnedUnit", ""),
            }

            lote.append((
                "Entrenamiento",
                tipo_original,
                nombre_tipo(tipo_original),
                elem.attrib.get("sourceName", ""),
                elem.attrib.get("durationUnit", ""),
                "min",
                elem.attrib.get("duration", ""),
                duracion,
                str(round(duracion, 6)) if duracion is not None else "",
                fecha_a_texto(inicio_dt),
                fecha_a_texto(fin_dt),
                duracion_minutos(inicio_dt, fin_dt) if duracion is None else duracion,
                fecha_a_texto(creacion_dt),
                distancia_km,
                energia_kcal,
                json.dumps(extra, ensure_ascii=False)
            ))
            contadores["Entrenamiento"] += 1

        elif tag == "ActivitySummary":
            fecha = elem.attrib.get("dateComponents", "")
            inicio_dt = parse_fecha_apple(fecha)

            if not dentro_de_fechas(inicio_dt, fecha_inicio, fecha_fin_exclusiva):
                elem.clear()
                continue

            extra = dict(elem.attrib)
            lote.append((
                "Resumen actividad",
                "HKActivitySummary",
                "Resumen actividad diaria",
                "",
                "",
                "",
                "",
                None,
                "",
                fecha_a_texto(inicio_dt),
                "",
                None,
                "",
                None,
                None,
                json.dumps(extra, ensure_ascii=False)
            ))
            contadores["Resumen actividad"] += 1

        if len(lote) >= INSERT_CHUNK:
            insertar_lote(conn, lote)
            lote.clear()

        elem.clear()

    if lote:
        insertar_lote(conn, lote)

    return dict(contadores)


def periodo_sql(detalle):
    if detalle == "dia":
        return "substr(inicio, 1, 10)"
    if detalle == "hora":
        return "substr(inicio, 1, 13) || ':00'"
    if detalle == "min":
        return "substr(inicio, 1, 16)"
    raise ValueError("periodo_sql no aplica para detalle=todos")


def nombre_hoja_seguro(nombre):
    nombre = re.sub(r"[\[\]\:\*\?\/\\]", " ", nombre)
    nombre = re.sub(r"\s+", " ", nombre).strip()
    return nombre or "Hoja"


def hoja_unica(base, usadas, sufijo=""):
    base = nombre_hoja_seguro(base)
    if sufijo:
        max_base = 31 - len(sufijo)
        candidato = (base[:max_base] + sufijo)[:31]
    else:
        candidato = base[:31]

    original = candidato
    i = 2
    while candidato in usadas:
        cola = f"_{i}"
        candidato = (original[:31 - len(cola)] + cola)[:31]
        i += 1

    usadas.add(candidato)
    return candidato



# =========================
# Formato Excel
# =========================

# Estilo pensado para archivos grandes: se formatean cabeceras, pestañas, filtros,
# congelación y anchos de columna. No se aplican bordes a millones de celdas para
# no hacer el Excel enorme ni lentísimo.

THEME = {
    "azul": "1F4E78",
    "azul_claro": "D9EAF7",
    "verde": "70AD47",
    "naranja": "ED7D31",
    "morado": "8064A2",
    "gris": "F2F2F2",
    "gris_texto": "404040",
    "blanco": "FFFFFF",
    "borde": "B7B7B7",
}

TAB_COLOR_BY_CLASS = {
    "Registro": THEME["azul"],
    "Categoría": THEME["verde"],
    "Entrenamiento": THEME["naranja"],
    "Resumen actividad": THEME["morado"],
}


def _style_objects():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=THEME["borde"])
    return {
        "title_font": Font(name="Calibri", size=16, bold=True, color=THEME["azul"]),
        "subtitle_font": Font(name="Calibri", size=10, italic=True, color=THEME["gris_texto"]),
        "section_font": Font(name="Calibri", size=11, bold=True, color=THEME["blanco"]),
        "header_font": Font(name="Calibri", size=10, bold=True, color=THEME["blanco"]),
        "normal_font": Font(name="Calibri", size=10, color="000000"),
        "title_fill": PatternFill("solid", fgColor=THEME["blanco"]),
        "section_fill": PatternFill("solid", fgColor=THEME["azul"]),
        "header_fill": PatternFill("solid", fgColor=THEME["azul"]),
        "meta_header_fill": PatternFill("solid", fgColor=THEME["azul_claro"]),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
    }


def celda_estilo(ws, valor, tipo="data"):
    """Devuelve una celda write-only con estilo para cabeceras/resumen."""
    from openpyxl.cell import WriteOnlyCell

    styles = _style_objects()
    cell = WriteOnlyCell(ws, value=valor)

    if tipo == "title":
        cell.font = styles["title_font"]
        cell.fill = styles["title_fill"]
        cell.alignment = styles["left"]
    elif tipo == "subtitle":
        cell.font = styles["subtitle_font"]
        cell.alignment = styles["wrap"]
    elif tipo == "section":
        cell.font = styles["section_font"]
        cell.fill = styles["section_fill"]
        cell.alignment = styles["left"]
        cell.border = styles["border"]
    elif tipo == "header":
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    elif tipo == "meta_header":
        cell.font = styles["normal_font"]
        cell.fill = styles["meta_header_fill"]
        cell.alignment = styles["left"]
        cell.border = styles["border"]
    else:
        cell.font = styles["normal_font"]
        cell.alignment = styles["left"]

    return cell


def fila_estilo(ws, valores, tipo="data"):
    if tipo == "data":
        # Mantener datos sin estilo celda a celda mejora mucho el rendimiento en exports grandes.
        return list(valores)
    return [celda_estilo(ws, v, tipo) for v in valores]



def celda_fc(ws, valor, estilo="normal"):
    """Celda específica para la hoja Resumen_Frecuencia cardiaca."""
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cell = WriteOnlyCell(ws, value=valor)
    thin = Side(style="thin", color=THEME["borde"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if estilo == "titulo":
        cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    elif estilo == "banda_azul":
        cell.font = Font(name="Calibri", size=10, bold=True, color=THEME["blanco"])
        cell.fill = PatternFill("solid", fgColor=THEME["azul"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    elif estilo == "cabecera_verde":
        cell.font = Font(name="Calibri", size=10, bold=True, color=THEME["blanco"])
        cell.fill = PatternFill("solid", fgColor="0B3D1A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    elif estilo == "etiqueta_verde":
        cell.font = Font(name="Calibri", size=10, bold=False, color=THEME["blanco"])
        cell.fill = PatternFill("solid", fgColor="0B3D1A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    elif estilo == "dato":
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    elif estilo == "dato_destacado":
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    elif estilo == "blanco_borde":
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    else:
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    return cell


def _valor_resumen_fc(valor):
    if valor is None:
        return ""
    if isinstance(valor, float):
        return round(valor, 3)
    return valor


def resumen_frecuencia_cardiaca_stats(conn, detalle, tipo_original):
    """
    Calcula los indicadores del resumen cardiaco a partir de los mismos datos filtrados:
    - Muestras: suma de mediciones del periodo seleccionado.
    - Promedio: promedio de las columnas minimo/media/maximo por intervalo.
    - MIN/MAX: extremos de esas columnas por intervalo.

    Para --detalle todos, cada registro se trata como un intervalo individual.
    """
    if detalle == "todos":
        subquery = """
            SELECT
                COUNT(*) AS mediciones,
                MIN(valor_num) AS minimo,
                AVG(valor_num) AS media,
                MAX(valor_num) AS maximo
            FROM salud
            WHERE tipo_original = ?
              AND clase = 'Registro'
              AND valor_num IS NOT NULL
            GROUP BY id
        """
    else:
        p = periodo_sql(detalle)
        subquery = f"""
            SELECT
                {p} AS periodo,
                fuente,
                unidad,
                COUNT(*) AS mediciones,
                MIN(valor_num) AS minimo,
                AVG(valor_num) AS media,
                MAX(valor_num) AS maximo
            FROM salud
            WHERE tipo_original = ?
              AND clase = 'Registro'
              AND valor_num IS NOT NULL
            GROUP BY periodo, fuente, unidad
        """

    row = conn.execute(f"""
        SELECT
            COALESCE(SUM(mediciones), 0) AS muestras,
            AVG(minimo) AS promedio_minimo,
            AVG(media) AS promedio_media,
            AVG(maximo) AS promedio_maximo,
            MIN(minimo) AS min_minimo,
            MIN(media) AS min_media,
            MIN(maximo) AS min_maximo,
            MAX(minimo) AS max_minimo,
            MAX(media) AS max_media,
            MAX(maximo) AS max_maximo
        FROM ({subquery}) grupos
    """, (tipo_original,)).fetchone()

    if row is None:
        return {
            "muestras": 0,
            "promedio_minimo": None,
            "promedio_media": None,
            "promedio_maximo": None,
            "min_minimo": None,
            "min_media": None,
            "min_maximo": None,
            "max_minimo": None,
            "max_media": None,
            "max_maximo": None,
        }

    keys = [
        "muestras",
        "promedio_minimo",
        "promedio_media",
        "promedio_maximo",
        "min_minimo",
        "min_media",
        "min_maximo",
        "max_minimo",
        "max_media",
        "max_maximo",
    ]
    return dict(zip(keys, row))


def escribir_resumen_frecuencia_cardiaca(wb, usadas, conn, detalle):
    """
    Crea una hoja tipo dashboard para los tres indicadores cardiacos principales,
    siguiendo la estructura del ejemplo aportado:
    - Frecuencia cardiaca
    - Frecuencia cardiaca caminando
    - Frecuencia cardiaca en reposo
    """
    bloques = [
        ("Frecuencia cardiaca", "HKQuantityTypeIdentifierHeartRate", 1),
        ("Frecuencia cardiaca caminando", "HKQuantityTypeIdentifierWalkingHeartRateAverage", 7),
        ("Frecuencia cardiaca en reposo", "HKQuantityTypeIdentifierRestingHeartRate", 13),
    ]

    ws = wb.create_sheet(title=hoja_unica("Resumen_Frecuencia cardiaca", usadas))
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = THEME["azul"]
    ws.freeze_panes = "A4"

    from openpyxl.utils import get_column_letter
    for col in range(1, 17):
        letter = get_column_letter(col)
        if col in {5, 6, 11, 12}:
            ws.column_dimensions[letter].width = 4
        else:
            ws.column_dimensions[letter].width = 13

    stats_por_tipo = {
        tipo_original: resumen_frecuencia_cardiaca_stats(conn, detalle, tipo_original)
        for _, tipo_original, _ in bloques
    }

    matriz = [[celda_fc(ws, "", "normal") for _ in range(16)] for _ in range(10)]

    for titulo, tipo_original, col_ini in bloques:
        matriz[0][col_ini - 1] = celda_fc(ws, titulo, "titulo")

    for titulo, tipo_original, col_ini in bloques:
        s = stats_por_tipo[tipo_original]
        i = col_ini - 1

        matriz[2][i + 0] = celda_fc(ws, "", "banda_azul")
        matriz[2][i + 1] = celda_fc(ws, "Promedio", "banda_azul")
        matriz[2][i + 2] = celda_fc(ws, "", "banda_azul")
        matriz[2][i + 3] = celda_fc(ws, "", "banda_azul")

        for offset, header in enumerate(["Muestras", "Mínimo", "Media", "Máximo"]):
            matriz[3][i + offset] = celda_fc(ws, header, "cabecera_verde")

        valores_promedio = [
            int(s["muestras"] or 0),
            _valor_resumen_fc(s["promedio_minimo"]),
            _valor_resumen_fc(s["promedio_media"]),
            _valor_resumen_fc(s["promedio_maximo"]),
        ]
        for offset, value in enumerate(valores_promedio):
            matriz[4][i + offset] = celda_fc(ws, value, "dato")

        matriz[6][i + 0] = celda_fc(ws, "", "banda_azul")
        matriz[6][i + 1] = celda_fc(ws, "Máximos y mínimos", "banda_azul")
        matriz[6][i + 2] = celda_fc(ws, "", "banda_azul")
        matriz[6][i + 3] = celda_fc(ws, "", "banda_azul")

        matriz[7][i + 0] = celda_fc(ws, "", "blanco_borde")
        for offset, header in enumerate(["Mínimo", "Media", "Máximo"], start=1):
            matriz[7][i + offset] = celda_fc(ws, header, "cabecera_verde")

        matriz[8][i + 0] = celda_fc(ws, "MIN", "etiqueta_verde")
        min_values = [
            _valor_resumen_fc(s["min_minimo"]),
            _valor_resumen_fc(s["min_media"]),
            _valor_resumen_fc(s["min_maximo"]),
        ]
        for offset, value in enumerate(min_values, start=1):
            matriz[8][i + offset] = celda_fc(ws, value, "dato_destacado" if offset in {1, 3} else "dato")

        matriz[9][i + 0] = celda_fc(ws, "MAX", "etiqueta_verde")
        max_values = [
            _valor_resumen_fc(s["max_minimo"]),
            _valor_resumen_fc(s["max_media"]),
            _valor_resumen_fc(s["max_maximo"]),
        ]
        for offset, value in enumerate(max_values, start=1):
            matriz[9][i + offset] = celda_fc(ws, value, "dato" if offset == 2 else "dato_destacado")

    for fila in matriz:
        ws.append(fila)



def ancho_columna(nombre_columna):
    n = (nombre_columna or "").lower()
    if n in {"periodo", "inicio", "fin", "creacion", "primera_fecha", "ultima_fecha", "fecha_hora"}:
        return 21
    if n in {"fecha"}:
        return 13
    if n in {"hora"}:
        return 9
    if "tipo_original" in n:
        return 42
    if n in {"tipo", "clase", "estado", "fuente", "fuente_principal"}:
        return 28
    if "unidad" in n:
        return 16
    if n in {"extra"}:
        return 50
    if "valor_texto" in n:
        return 24
    if any(k in n for k in ["mediciones", "registros", "entrenos"]):
        return 13
    if any(k in n for k in ["min", "max", "media", "suma", "total", "kcal", "km", "valor_num"]):
        return 15
    return min(max(len(str(nombre_columna)) + 4, 12), 26)


def preparar_hoja(ws, columnas=None, tab_color=None, freeze="A2", filtro=True):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.zoomScale = 90

    if columnas:
        for i, col in enumerate(columnas, start=1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = ancho_columna(col)


TABLE_NAMES_USED = set()


def nombre_tabla_unico(nombre_hoja):
    """Nombre válido y único para tablas de Excel."""
    base = re.sub(r"[^A-Za-z0-9_]", "_", nombre_hoja or "Tabla")
    base = re.sub(r"_+", "_", base).strip("_") or "Tabla"
    if not re.match(r"^[A-Za-z_]", base):
        base = f"T_{base}"
    # Evitar nombres demasiado largos y posibles referencias de celda.
    base = base[:200]
    candidato = base
    i = 2
    while candidato in TABLE_NAMES_USED:
        candidato = f"{base[:190]}_{i}"
        i += 1
    TABLE_NAMES_USED.add(candidato)
    return candidato


def finalizar_filtro(ws, columnas, filas_datos):
    if not columnas:
        return
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

    last_col = get_column_letter(len(columnas))
    last_row = max(filas_datos + 1, 1)
    ref = f"A1:{last_col}{last_row}"

    # AutoFilter clásico + tabla con bandas: así el formato es visible aunque el libro
    # se haya escrito en modo streaming para aguantar exports grandes.
    ws.auto_filter.ref = ref

    if filas_datos <= 0:
        return

    try:
        tabla = Table(displayName=nombre_tabla_unico(ws.title), ref=ref)
        tabla.tableColumns = [
            TableColumn(id=i + 1, name=str(col) if col is not None else f"Columna_{i+1}")
            for i, col in enumerate(columnas)
        ]
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            ws.add_table(tabla)
    except Exception as e:
        # No bloqueamos la exportación por un problema cosmético de tabla.
        print(f"Aviso: no se pudo aplicar estilo de tabla en hoja {ws.title}: {e}")


def write_query_en_hojas(wb, conn, usadas, nombre_base, columnas, query, params=(), clase=None):
    cur = conn.execute(query, params)

    parte = 1
    filas_en_hoja = 0
    ws = None
    tab_color = TAB_COLOR_BY_CLASS.get(clase, THEME["azul"])

    def nueva_hoja():
        nonlocal parte, filas_en_hoja
        sufijo = "" if parte == 1 else f"_{parte}"
        titulo = hoja_unica(nombre_base, usadas, sufijo=sufijo)
        hoja = wb.create_sheet(title=titulo)
        preparar_hoja(hoja, columnas=columnas, tab_color=tab_color, freeze="A2")
        hoja.append(fila_estilo(hoja, columnas, "header"))
        hoja.row_dimensions[1].height = 22
        filas_en_hoja = 0
        parte += 1
        return hoja

    ws = nueva_hoja()

    for row in cur:
        if filas_en_hoja >= EXCEL_DATA_ROWS:
            finalizar_filtro(ws, columnas, filas_en_hoja)
            ws = nueva_hoja()

        ws.append(fila_estilo(ws, row, "data"))
        filas_en_hoja += 1

    finalizar_filtro(ws, columnas, filas_en_hoja)


def escribir_resumen(wb, usadas, conn, xml_path, detalle, fecha_inicio_txt, fecha_fin_txt, contadores):
    ws = wb.create_sheet(title=hoja_unica("Resumen", usadas))
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = THEME["azul"]
    ws.freeze_panes = "A8"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 16

    ws.append([celda_estilo(ws, "Apple Health Export", "title")])
    ws.append([celda_estilo(ws, "Resumen de la exportación generada desde export.xml", "subtitle")])
    ws.append([])

    ws.append([celda_estilo(ws, "Configuración", "section"), ""])
    ws.append(fila_estilo(ws, ["Campo", "Valor"], "header"))
    ws.append(["Archivo XML", str(xml_path)])
    ws.append(["Detalle", detalle])
    ws.append(["Fecha inicio", fecha_inicio_txt or "sin límite"])
    ws.append(["Fecha fin", fecha_fin_txt or "sin límite"])
    ws.append(["Nota fecha fin", "La fecha fin se interpreta como día completo incluido."])
    ws.append(["Límite exacto Excel", f"{EXCEL_MAX_ROWS} filas por hoja"])
    ws.append(["Filas máximas de datos por hoja", str(EXCEL_DATA_ROWS)])
    ws.append([])

    ws.append([celda_estilo(ws, "Registros importados por clase", "section"), ""])
    ws.append(fila_estilo(ws, ["Clase", "Registros importados"], "header"))
    for clase, n in sorted(contadores.items()):
        ws.append([clase, n])
    ws.append([])

    ws.append([celda_estilo(ws, "Tipos detectados", "section"), "", "", "", "", ""])
    columnas = ["Tipo", "Tipo original", "Clase", "Registros", "Primera fecha", "Última fecha"]
    ws.append(fila_estilo(ws, columnas, "header"))

    filas_tipo = 0
    for row in conn.execute("""
        SELECT tipo, tipo_original, clase, COUNT(*), MIN(inicio), MAX(inicio)
        FROM salud
        GROUP BY tipo_original, tipo, clase
        ORDER BY tipo
    """):
        ws.append(list(row))
        filas_tipo += 1

    # El filtro se aplica al bloque principal de tipos. Aunque hay bloques anteriores,
    # sigue siendo útil para localizar rápidamente cada tipo de dato.
    from openpyxl.utils import get_column_letter
    inicio_tabla = 16 + len(contadores) + 4
    fin_tabla = inicio_tabla + filas_tipo
    ws.auto_filter.ref = f"A{inicio_tabla}:{get_column_letter(len(columnas))}{fin_tabla}"


def exportar_excel(conn, salida, detalle, xml_path, fecha_inicio_txt, fecha_fin_txt, contadores):
    wb = Workbook(write_only=True)
    usadas = set()

    escribir_resumen(
        wb=wb,
        usadas=usadas,
        conn=conn,
        xml_path=xml_path,
        detalle=detalle,
        fecha_inicio_txt=fecha_inicio_txt,
        fecha_fin_txt=fecha_fin_txt,
        contadores=contadores,
    )

    escribir_resumen_frecuencia_cardiaca(
        wb=wb,
        usadas=usadas,
        conn=conn,
        detalle=detalle,
    )

    tipos = list(conn.execute("""
        SELECT tipo_original, tipo, clase, COUNT(*)
        FROM salud
        GROUP BY tipo_original, tipo, clase
        ORDER BY tipo
    """))

    if detalle == "todos":
        columnas = [
            "clase", "tipo", "tipo_original", "fuente",
            "inicio", "fin", "duracion_min",
            "valor_num", "valor_texto", "unidad", "unidad_original",
            "distancia_km", "energia_kcal",
            "creacion", "extra"
        ]

        for tipo_original, tipo, clase, _ in tipos:
            query = """
                SELECT
                    clase, tipo, tipo_original, fuente,
                    inicio, fin, duracion_min,
                    valor_num, valor_texto, unidad, unidad_original,
                    distancia_km, energia_kcal,
                    creacion, extra
                FROM salud
                WHERE tipo_original = ? AND tipo = ? AND clase = ?
                ORDER BY inicio
            """
            write_query_en_hojas(wb, conn, usadas, tipo, columnas, query, (tipo_original, tipo, clase), clase=clase)

    else:
        p = periodo_sql(detalle)

        for tipo_original, tipo, clase, _ in tipos:
            if clase == "Registro":
                columnas = [
                    "periodo", "tipo", "fuente", "unidad",
                    "mediciones", "minimo", "media", "maximo", "suma",
                    "primera_fecha", "ultima_fecha"
                ]
                query = f"""
                    SELECT
                        {p} AS periodo,
                        tipo,
                        fuente,
                        unidad,
                        COUNT(*) AS mediciones,
                        ROUND(MIN(valor_num), 6) AS minimo,
                        ROUND(AVG(valor_num), 6) AS media,
                        ROUND(MAX(valor_num), 6) AS maximo,
                        ROUND(SUM(valor_num), 6) AS suma,
                        MIN(inicio) AS primera_fecha,
                        MAX(inicio) AS ultima_fecha
                    FROM salud
                    WHERE tipo_original = ? AND tipo = ? AND clase = ?
                      AND valor_num IS NOT NULL
                    GROUP BY periodo, tipo, fuente, unidad
                    ORDER BY periodo
                """
                write_query_en_hojas(wb, conn, usadas, tipo, columnas, query, (tipo_original, tipo, clase), clase=clase)

            elif clase == "Categoría":
                columnas = [
                    "periodo", "tipo", "fuente", "estado",
                    "registros", "minutos_total", "minutos_media_registro",
                    "primera_fecha", "ultima_fecha"
                ]
                query = f"""
                    SELECT
                        {p} AS periodo,
                        tipo,
                        fuente,
                        valor_texto AS estado,
                        COUNT(*) AS registros,
                        ROUND(SUM(duracion_min), 6) AS minutos_total,
                        ROUND(AVG(duracion_min), 6) AS minutos_media_registro,
                        MIN(inicio) AS primera_fecha,
                        MAX(inicio) AS ultima_fecha
                    FROM salud
                    WHERE tipo_original = ? AND tipo = ? AND clase = ?
                    GROUP BY periodo, tipo, fuente, estado
                    ORDER BY periodo
                """
                write_query_en_hojas(wb, conn, usadas, tipo, columnas, query, (tipo_original, tipo, clase), clase=clase)

            elif clase == "Entrenamiento":
                columnas = [
                    "periodo", "tipo", "fuente",
                    "entrenos", "minutos_total", "minutos_media",
                    "distancia_km_total", "energia_kcal_total",
                    "primera_fecha", "ultima_fecha"
                ]
                query = f"""
                    SELECT
                        {p} AS periodo,
                        tipo,
                        fuente,
                        COUNT(*) AS entrenos,
                        ROUND(SUM(duracion_min), 6) AS minutos_total,
                        ROUND(AVG(duracion_min), 6) AS minutos_media,
                        ROUND(SUM(distancia_km), 6) AS distancia_km_total,
                        ROUND(SUM(energia_kcal), 6) AS energia_kcal_total,
                        MIN(inicio) AS primera_fecha,
                        MAX(inicio) AS ultima_fecha
                    FROM salud
                    WHERE tipo_original = ? AND tipo = ? AND clase = ?
                    GROUP BY periodo, tipo, fuente
                    ORDER BY periodo
                """
                write_query_en_hojas(wb, conn, usadas, tipo, columnas, query, (tipo_original, tipo, clase), clase=clase)

            else:
                columnas = [
                    "clase", "tipo", "tipo_original", "inicio", "valor_texto", "extra"
                ]
                query = """
                    SELECT clase, tipo, tipo_original, inicio, valor_texto, extra
                    FROM salud
                    WHERE tipo_original = ? AND tipo = ? AND clase = ?
                    ORDER BY inicio
                """
                write_query_en_hojas(wb, conn, usadas, tipo, columnas, query, (tipo_original, tipo, clase), clase=clase)

    wb.save(salida)

def main():
    parser = argparse.ArgumentParser(
        description="Exporta Apple Health export.xml a Excel, con una hoja por tipo de dato, detalle por día/hora/minuto/todos y resumen de frecuencia cardiaca."
    )
    parser.add_argument("--xml", default=None, help="Ruta opcional a export.xml o a la carpeta que lo contiene. Si no se indica, se busca en subcarpetas.")
    parser.add_argument("--inicio", default=None, help="Fecha inicial incluida. Formato YYYY-MM-DD.")
    parser.add_argument("--fin", default=None, help="Fecha final incluida. Formato YYYY-MM-DD.")
    parser.add_argument("--detalle", default="hora", help="dia, hora, min o todos.")
    parser.add_argument("--salida", default=None, help="Nombre del Excel de salida.")
    parser.add_argument("--db", default="apple_health_temp.sqlite", help="SQLite temporal.")
    parser.add_argument("--version", action="version", version=VERSION)

    args = parser.parse_args()

    detalle = normalizar_detalle(args.detalle)
    fecha_inicio = parse_fecha_arg(args.inicio, "--inicio")
    fecha_fin = parse_fecha_arg(args.fin, "--fin")
    fecha_fin_exclusiva = fecha_fin + timedelta(days=1) if fecha_fin else None

    xml_path = buscar_export_xml(args.xml)

    if args.salida:
        salida = args.salida
    else:
        sufijo_fechas = ""
        if args.inicio or args.fin:
            sufijo_fechas = f"_{args.inicio or 'inicio'}_a_{args.fin or 'fin'}"
        salida = f"apple_health_{detalle}{sufijo_fechas}.xlsx"

    db_path = Path(args.db)

    print(f"Versión script: {VERSION}")
    print(f"XML: {xml_path}")
    print(f"Detalle: {detalle}")
    print(f"Fecha inicio: {args.inicio or 'sin límite'}")
    print(f"Fecha fin: {args.fin or 'sin límite'}")
    print(f"Salida: {salida}")

    conn = crear_db(db_path)
    contadores = procesar_xml(xml_path, conn, fecha_inicio, fecha_fin_exclusiva)

    total = conn.execute("SELECT COUNT(*) FROM salud").fetchone()[0]
    tipos = conn.execute("SELECT COUNT(DISTINCT tipo_original || '|' || tipo || '|' || clase) FROM salud").fetchone()[0]

    print(f"Registros importados: {total:,}")
    print(f"Tipos/hojas base detectados: {tipos:,}")
    print("Creando Excel...")

    exportar_excel(
        conn=conn,
        salida=salida,
        detalle=detalle,
        xml_path=xml_path,
        fecha_inicio_txt=args.inicio,
        fecha_fin_txt=args.fin,
        contadores=contadores,
    )

    conn.close()

    print("")
    print("Listo.")
    print(f"Excel creado con formato: {salida}")
    print(f"SQLite temporal: {db_path}")
    print("")
    print("Aviso: en detalle dia/hora/min, los datos con duración se agrupan por la fecha/hora de inicio del registro.")


if __name__ == "__main__":
    main()
